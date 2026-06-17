#!/usr/bin/env python3
"""
generate_iif.py
Generates 4 QuickBooks IIF payroll journal entry files for Cornerstones
from a Paycom source Excel workbook.

Usage:
    python3 generate_iif.py \
        --source "Source 1 FY26 - PD 20260213 Payroll - Done.xlsx" \
        --class-ref "Reference 1 Class Table.xlsx" \
        --pay-date "2/13/2026" \
        --pay-period 17 \
        --fiscal-year 26 \
        --output-dir "/path/to/output"
"""

import argparse
import os
import sys
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl --break-system-packages -q")
    import openpyxl


# ---------------------------------------------------------------------------
# IIF constants — QB account paths (Cornerstones chart of accounts)
# ---------------------------------------------------------------------------
ACCT_SALARY       = "DIRECT PROGRAM EXPENSES:Personnel:Salaries:Direct Salaries (Full Time)"
ACCT_TAX          = "DIRECT PROGRAM EXPENSES:Personnel:Fringe Benefits:Employee Payroll Taxes"
ACCT_401K         = "DIRECT PROGRAM EXPENSES:Personnel:Fringe Benefits:401K"
ACCT_FEES         = "DIRECT PROGRAM EXPENSES:Personnel:Fringe Benefits:Benefits Fees-Admin"
ACCT_HEALTH       = "DIRECT PROGRAM EXPENSES:Personnel:Fringe Benefits:Health"
ACCT_DISABILITY   = "DIRECT PROGRAM EXPENSES:Personnel:Fringe Benefits:Disability & Life Ins."
ACCT_CASH         = "Cash:Access Bank (HEADER):Atlantic Union Operations Chkng"
ACCT_SUSPENSE          = "Payroll Related Payable:Payroll Suspense -Clearing Acct"
ACCT_SUSPENSE_MANUAL   = "Payroll Related Payable:Payroll Suspense -Clearing Acct"
ACCT_GARNISHMENT  = "Payroll Related Payable:Garnishment Withheld"
ACCT_CELL_PHONE   = "Payroll Related Payable:Cell Phone Reimbursement"
ACCT_EXPENSE_REIMB= "Payroll Related Payable:Expense Reimbursement"
ACCT_MILEAGE      = "Payroll Related Payable:Mileage Reimbursement"
ACCT_401K_PAYABLE = "Payroll Related Payable:401K Contributions Payable"
ACCT_FSA          = "Payroll Related Payable:FSA Withheld"
ACCT_DENTAL       = "Payroll Related Payable:Dental Premiums Withheld"
ACCT_LEGAL        = "Payroll Related Payable:Legal Resources Withheld"
ACCT_LLC_TUITION  = "Payroll Related Payable:LLC Tuition Fees Witheld"   # QB spells it "Witheld" (one h)
CLASS_UNALLOCATED = "901 Unallocated Personnel Costs"

# Paycom account code groupings for Sheet0 aggregation
GARNISHMENT_DESCS = {
    "Payable - Garnishment-1 $", "Payable - Child Support #1 $",
    "Payable - Child Support #2 $", "Payable - Child Support Order #3",
    "Payable - Child Support Order #4",
    # Add more child support / garnishment codes as they appear
}
FSA_DESCS   = {"Payable - Flexible Spending Med", "Payable - Kaiser Flex F Plan"}
DENTAL_METLIFE_DESCS = {"Payable - Metlife Dental Plan", "Payable - Metlife Dental"}
VISION_UHC_DESCS     = {"Payable - Vision Plan", "Payable - UHC Vision"}
LEGAL_DESCS          = {"Payable - Legal Resources Post-", "Payable - Legal Resources Withheld"}
HEALTH_SENTARA_DESCS = {"9012.1", "Employer Kaiser", "Health - Sentara", "Health - Kaiser"}  # matched by account code prefix
DISABILITY_AD_DESCS  = {"Payable - AD&D Post-Tax", "Payable - Voluntary EE Life Pla",
                         "Payable - Voluntary SP Life Pla", "Payable - Voluntary CH Life Pla",
                         "Payable - Child AD&D", "Payable - Spouse AD&D",
                         "Payable - Employer Paid Life an"}
DISABILITY_LTD_DESCS = {"Payable - Long Term Disability", "Payable - Short Term disability"}
CELL_PHONE_DESCS     = {"6421", "Cell Phone"}
MILEAGE_DESCS        = {"6420", "Mileage"}
EXPENSE_REIMB_DESCS  = {"6422", "Expense Reimb"}
EE_401K_DESCS        = {"Payable - 401K %", "Payable - 401K$",
                         "Payable - 401K Loan Repayment 1", "Payable - 401K Loan Repayment 2",
                         "Payable - 401K Loan Repayment 3"}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def load_final_check_table(path):
    """
    Load Reference 2 Class Table FINAL CHECK.xlsx.
    Returns a set of valid QB class paths (stripped, lowercased for comparison).
    The workbook has a single column of full class paths starting in row 3
    (row 1 is a header 'Class Masterlist', row 2 is blank).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    valid = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue  # skip header and blank row
        val = row[0]
        if val and str(val).strip():
            valid.add(str(val).strip())
    wb.close()
    return valid


def verify_classes_against_final_check(iif_paths, final_check_table):
    """
    Read all generated IIF files, extract every CLASS value (column index 6,
    0-based, in tab-delimited TRNS/SPL rows), and verify each one exists in
    final_check_table.  Blank class values and '901 Unallocated Personnel Costs'
    (an internal QB class used for offset rows) are always allowed.

    Returns a list of (file, class_value) tuples for any missing classes.
    """
    ALWAYS_ALLOWED = {"", " ", "901 Unallocated Personnel Costs"}
    missing = []
    for path in iif_paths:
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.rstrip("\n\r")
                if not (line.startswith("TRNS") or line.startswith("SPL")):
                    continue
                parts = line.split("\t")
                if len(parts) < 7:
                    continue
                cls = parts[6].strip()
                if cls in ALWAYS_ALLOWED:
                    continue
                if cls not in final_check_table:
                    missing.append((os.path.basename(path), cls))
    return missing


def fmt_amount(value):
    """Format a number to 2 decimal places."""
    return f"{value:.2f}"


def parse_pay_date(pay_date_str):
    """Parse pay date string like '2/13/2026' into a datetime."""
    for fmt in ("%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(pay_date_str, fmt)
        except ValueError:
            pass
    raise ValueError(f"Cannot parse pay date: {pay_date_str}")


def iif_date(dt):
    """Format date for IIF: M/D/YYYY (no leading zeros)."""
    return f"{dt.month}/{dt.day}/{dt.year}"


def file_date_label(dt):
    """Format date for file name: DD Mon (e.g., 13 Feb)."""
    return dt.strftime("%-d %b")


def docnum_prp(dt):
    """PRP02.13.26 format."""
    return f"PRP{dt.month:02d}.{dt.day:02d}.{str(dt.year)[2:]}"


def docnum_tax(dt):
    """Tax2.13.26 format (no leading zero on month)."""
    return f"Tax{dt.month}.{dt.day}.{str(dt.year)[2:]}"


def docnum_401k(dt):
    """401K.021326 format."""
    return f"401K.{dt.month:02d}{dt.day:02d}{str(dt.year)[2:]}"


def docnum_fee(dt):
    """Fee.021326 format."""
    return f"Fee.{dt.month:02d}{dt.day:02d}{str(dt.year)[2:]}"


def memo_prefix(fy, pr):
    """Build memo prefix like 'FY26.PR.17'."""
    return f"FY{fy}.PR.{pr}"


def build_iif_header():
    """Return the 3-line IIF header."""
    trns_cols = ["TRNSID", "TRNSTYPE", "DATE", "ACCNT", "NAME", "CLASS", "AMOUNT", "DOCNUM", "MEMO"]
    spl_cols  = ["SPLID",  "TRNSTYPE", "DATE", "ACCNT", "NAME", "CLASS", "AMOUNT", "DOCNUM", "MEMO"]
    hdr  = "\t".join(["!TRNS"] + trns_cols)
    spl  = "\t".join(["!SPL"]  + spl_cols)
    end  = "!ENDTRNS"
    return hdr + "\n" + spl + "\n" + end


def build_row(row_type, trnstype, date_str, accnt, name, cls, amount, docnum, memo):
    """Build a single tab-delimited IIF row."""
    # Warn whenever a NAME value is written — the name must exist in QuickBooks
    # or the import will fail. Prompt the user to verify before importing.
    if name and name.strip():
        print(f"  ⚠ NAME used: '{name.strip()}' (account: {accnt}) — verify this name exists in QuickBooks before importing.")
    fields = [
        row_type,
        "",        # TRNSID / SPLID (always blank)
        trnstype,
        date_str,
        accnt,
        name,
        cls,
        fmt_amount(amount),
        docnum,
        memo,
    ]
    return "\t".join(fields)


def write_iif(path, header, rows):
    """Write the complete IIF file."""
    with open(path, "w", encoding="utf-8", newline="\r\n") as f:
        f.write(header + "\n")
        for row in rows:
            f.write(row + "\n")
        f.write("ENDTRNS\n")
    print(f"  ✓ Written: {path}")


# ---------------------------------------------------------------------------
# Class table loading
# ---------------------------------------------------------------------------

def load_class_table(source_wb, ref_path):
    """
    Build a dict: str(code) -> full_QB_class_path.

    Loads exclusively from Reference 1 Class Table.xlsx (Sheet1_FileA,
    col A = code, col B = QB path).  The 'class worksheet' tab inside the
    source workbook is intentionally ignored.
    """
    table = {}

    if ref_path and os.path.exists(ref_path):
        try:
            ref_wb = openpyxl.load_workbook(ref_path, read_only=True, data_only=True)
            ws = ref_wb["Sheet1_FileA"]
            for row in ws.iter_rows(values_only=True):
                key, val = row[0], row[1]
                if key is not None and val is not None:
                    table[str(key).strip()] = str(val).strip()
            ref_wb.close()
            print(f"  Loaded {len(table)} entries from Reference 1 Class Table.")
        except Exception as e:
            print(f"  Warning: could not load class reference: {e}")

    return table


def lookup_class(code, table):
    """Look up a QB class path by code, trying str and numeric forms."""
    key = str(code).strip()
    # Try exact match
    if key in table:
        return table[key]
    # Try stripping trailing .0 (floats stored as 32101.0)
    if key.endswith(".0"):
        stripped = key[:-2]
        if stripped in table:
            return table[stripped]
    # Try converting to int
    try:
        int_key = str(int(float(key)))
        if int_key in table:
            return table[int_key]
    except (ValueError, OverflowError):
        pass
    return None


# ---------------------------------------------------------------------------
# Source sheet parsing
# ---------------------------------------------------------------------------

def read_output_sheet(wb, sheet_name, class_table=None):
    """
    Read an 'Output N ... By Code' sheet.
    Returns list of (code_str, net_amount_float) pairs in original order,
    skipping header rows and totals.
    Data rows: col B = Labor Allocation Code, col C = Description, col D = Dr amount, col E = Cr adjustment.
    Net = Dr + Cr (Cr is stored as negative when present).

    Codes with no class mapping are included in the output but flagged with
    a WARNING — their amounts will be excluded from the computed totals,
    causing the pivot check to fail (which is the correct behaviour: the
    source file must be corrected before the IIF can be imported).
    """
    ws = wb[sheet_name]
    results = []
    for row in ws.iter_rows(values_only=True):
        code = row[1] if len(row) > 1 else None
        desc = str(row[2] or '').strip() if len(row) > 2 else ''
        dr   = row[3] if len(row) > 3 else None
        cr   = row[4] if len(row) > 4 else None

        if code is None:
            continue
        if isinstance(code, str) and code.strip().lower() in (
            "labor allocation code", "account description", "amt", "", "grand total"
        ):
            continue
        if isinstance(code, str) and code.strip() == "":
            continue

        dr_val = float(dr) if isinstance(dr, (int, float)) else 0.0
        cr_val = float(cr) if isinstance(cr, (int, float)) else 0.0
        net = dr_val + cr_val
        if net == 0.0:
            continue

        code_str = str(code).strip()

        # Flag codes that have no class mapping — these indicate either a missing
        # entry in the Reference Class Table or an encoding error in the Paycom export.
        if class_table is not None and lookup_class(code_str, class_table) is None:
            print(f"  WARNING: No class mapping for salary code '{code_str}'"
                  f" ({desc}) — ${net:,.2f} skipping."
                  f" (Possible encoding error in Paycom — verify in source file.)")

        results.append((code_str, net))

    return results


def read_pivot_check_all(wb):
    """
    Read ALL Output 1 balance entries from the Pivot Check sheet right side (columns G–N).

    This is the single authoritative source for balance entries.  Sheet0 is NOT used.
    The Pivot Check operator maintains the correct, pre-adjusted values for every line.

    Column layout (0-indexed):
      G [6]  = full account label
      H [7]  = account code (e.g. "6412", "9012.1")
      I [8]  = QB class code (blank for most lines)
      J [9]  = debit amount (positive, or blank)
      K [10] = credit amount (negative, or blank)
      L [11] = "Name:" label (only on 6410 Suspense lines)
      M [12] = name value   (only on 6410 Suspense lines)
      N [13] = "Memo: …"   (only on 6410 Suspense lines)

    Returns:
        balance_entries (dict)   — same keys used by generate_output1()
        pivot_adjustments (list) — [{name, amount, memo}, …] individual SPL entries
    """
    pc_sheet = None
    for candidate in ("Pivot Check", "Pivot Check JVV"):
        if candidate in wb.sheetnames:
            pc_sheet = candidate
            break
    if pc_sheet is None:
        print("  WARNING: No Pivot Check sheet found — balance entries will be zero.")
        return {
            "garnishment": 0.0, "cell_phone": 0.0, "expense_reimb": 0.0,
            "mileage": 0.0, "ee_401k": 0.0, "fsa": 0.0,
            "dental_metlife": 0.0, "vision_uhc": 0.0, "legal": 0.0,
            "health_sentara": 0.0, "disability_ad": 0.0, "disability_ltd": 0.0,
            "oss_manual": 0.0, "llc_tuition": 0.0,
        }, []

    ws = wb[pc_sheet]
    bal = {
        "garnishment": 0.0, "cell_phone": 0.0, "expense_reimb": 0.0,
        "mileage": 0.0, "ee_401k": 0.0, "fsa": 0.0,
        "dental_metlife": 0.0, "vision_uhc": 0.0, "legal": 0.0,
        "health_sentara": 0.0, "disability_ad": 0.0, "disability_ltd": 0.0,
        "oss_manual": 0.0, "llc_tuition": 0.0,
    }
    pivot_adjustments = []

    # Codes whose amounts are derived from other Output sheets or computed elsewhere
    SKIP_CODES = {
        '1022.1',   # cash plug — computed as balancing entry
        '9011.1',   # salary — summed from Output 1 Salary By Code sheet
        '9012.2',   # 401K match debit — equals output3_total
        '9012.4',   # fee debit — equals output4_total
        '9012.6',   # payroll tax debit — equals output2_total
        '9090',     # miscellaneous (not in chart of accounts for this JE)
        '2410',     # due to other parties (not mapped to Output 1)
    }

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if len(row) < 10:
            continue

        label = str(row[6] or '').strip()   # col G
        code  = str(row[7] or '').strip()   # col H
        debit  = row[9]                     # col J
        credit = row[10]                    # col K

        if not code or code == 'Account':
            continue

        # Resolve amount: credit (negative) takes priority over debit (positive)
        if credit is not None and isinstance(credit, (int, float)) and credit != 0:
            amount = float(credit)
        elif debit is not None and isinstance(debit, (int, float)) and debit != 0:
            amount = float(debit)
        else:
            amount = 0.0

        label_lc = label.lower()

        # ── 6410 Payroll Suspense lines ──────────────────────────────────────
        if code == '6410':
            name_label = str(row[11] or '').strip() if len(row) > 11 else ''
            name_val   = str(row[12] or '').strip() if len(row) > 12 else ''
            memo_raw   = str(row[13] or '').strip() if len(row) > 13 else ''
            name = name_val.strip()
            memo = memo_raw[5:].strip() if memo_raw.lower().startswith('memo:') else memo_raw

            # Any 6410 line that has its own memo in column N becomes a separate
            # SPL line so the memo is preserved in the IIF.  Lines with no memo
            # are accumulated into oss_manual as before.
            if amount != 0.0:
                if memo:
                    pivot_adjustments.append({'name': name, 'amount': amount, 'memo': memo})
                    print(f"  6410 SPL line: name='{name}' | {amount:,.2f} | {memo}")
                else:
                    bal['oss_manual'] += amount
                    print(f"  OSS (Pivot Check): {amount:,.2f}")
            continue

        if code in SKIP_CODES or amount == 0.0:
            continue

        # ── Map account code + label to balance_entries key ──────────────────
        if code == '6412':
            bal['garnishment'] += amount

        elif code == '6413':
            # "Employee contributions" line → ee_401k
            # "Match" line → skip (handled by -output3_total in generate_output1)
            if 'employee' in label_lc:
                bal['ee_401k'] += amount

        elif code == '6415':
            bal['llc_tuition'] += amount

        elif code == '6417':
            bal['fsa'] += amount

        elif code == '6419':
            if 'vision' in label_lc:
                bal['vision_uhc'] += amount
            else:
                bal['dental_metlife'] += amount

        elif code == '6420':
            bal['mileage'] += amount          # positive debit

        elif code == '6421':
            bal['cell_phone'] += amount       # positive debit

        elif code == '6422':
            bal['expense_reimb'] += amount    # positive debit

        elif code == '6423':
            bal['legal'] += amount

        elif code == '9012.1':
            bal['health_sentara'] += amount   # covers Sentara, Kaiser, all 9012.1 health

        elif code == '9012.3':
            if 'ltd' in label_lc or 'long term' in label_lc:
                bal['disability_ltd'] += amount
            else:
                bal['disability_ad'] += amount

    return bal, pivot_adjustments


# ── Retained for backward-compatibility reference only (no longer called) ────
def aggregate_balance_entries(debit_agg, credit_agg, employer_memo_keys):
    """
    DEPRECATED — Sheet0 is no longer used as a source.
    Balance entries are now read exclusively from the Pivot Check sheet
    via read_pivot_check_all().  This function is retained only for reference.

    Employee withholdings (garnishments, 401K, FSA, dental, vision, legal, health, life, LTD)
    are credit entries in the payroll journal entry — they appear as negative/credit amounts.
    Reimbursements (cell phone, mileage, expense) are debit entries — positive amounts.

    Returns a dict with keys for each QB balance line.
    Negative values = credits (liabilities/withholdings); positive = debits (reimbursables).
    """
    bal = {
        "garnishment":      0.0,
        "cell_phone":       0.0,
        "expense_reimb":    0.0,
        "mileage":          0.0,
        "ee_401k":          0.0,
        "fsa":              0.0,
        "dental_metlife":   0.0,
        "vision_uhc":       0.0,
        "legal":            0.0,
        "health_sentara":   0.0,
        "disability_ad":    0.0,
        "disability_ltd":   0.0,
        "oss_manual":       0.0,   # Manual check cash payments (OSS etc.)
        "llc_tuition":      0.0,   # LLC Tuition Fees Withheld (6415)
    }

    # --- CREDIT entries (employee withholdings → credit/negative in the IIF) ---
    # We use credit_agg (amounts stored as positive) and negate them for the IIF.
    # Employer-paid memo counterparts are excluded via employer_memo_keys.
    for (code, desc), abs_amt in credit_agg.items():
        if (code, desc) in employer_memo_keys:
            continue   # exclude employer-paid memo entries (e.g., employer-paid STD/LTD/life)
        d = desc.lower()

        # Garnishments and child support withheld from employee paychecks
        if any(g in d for g in ['garnishment', 'child support', 'support order', 'support #']):
            bal['garnishment'] -= abs_amt

        # Employee 401K contributions (all types incl. loan repayments; exclude employer match)
        elif '401k' in d and 'match' not in d and 'employer' not in d:
            bal['ee_401k'] -= abs_amt

        # Flexible Spending Accounts
        elif 'flexible spending' in d:
            bal['fsa'] -= abs_amt

        # Dental — MetLife
        elif 'metlife' in d or 'dental plan' in d:
            bal['dental_metlife'] -= abs_amt

        # Vision Plan (UHC)
        elif 'vision plan' in d:
            bal['vision_uhc'] -= abs_amt

        # Legal Resources
        elif 'legal' in d:
            bal['legal'] -= abs_amt

        # Health — Sentara (employee-paid portion; "Memo" entries are employer-paid, excluded)
        elif 'sentara' in d and 'memo' not in d:
            bal['health_sentara'] -= abs_amt

        # Voluntary Life and AD&D (employee-elected coverage)
        elif any(k in d for k in ['voluntary', 'ad&d', 'child ad&d', 'spouse ad&d']):
            bal['disability_ad'] -= abs_amt

        # LTD — employee portion (employer LTD already filtered by employer_memo_keys)
        elif 'long term disability' in d:
            bal['disability_ltd'] -= abs_amt

        # LLC Tuition Fees Withheld (6415) — employee tuition deducted from paycheck
        elif 'llc tuition' in d or 'tuition' in d:
            bal['llc_tuition'] -= abs_amt

    # --- DEBIT entries (reimbursements owed to employees → debit/positive in the IIF) ---
    for (code, desc), amt in debit_agg.items():
        d = desc.lower()
        if 'cell phone' in d:
            bal['cell_phone'] += amt
        elif 'mileage' in d:
            bal['mileage'] += amt
        elif 'expense reimb' in d:
            bal['expense_reimb'] += amt

    # --- Manual check cash payments (OSS) → credit/negative in the IIF ---
    # These are net pay disbursements via manual/paper check rather than ACH.
    # In Sheet0 they appear as 'Cash - Manuals' credit entries (amount < 0).
    for (code, desc), abs_amt in credit_agg.items():
        if 'cash - manuals' in desc.lower():
            bal['oss_manual'] -= abs_amt   # credit (negative)

    return bal


def read_pivot_check_expected_amounts(wb):
    """
    Read the 'Pivot Check' sheet right-side table (columns G-K) and return
    the expected amounts for key account codes.  These are used after generation
    to validate that every computed total matches the source exactly.

    Returns a dict with keys:
      'cash'    → expected credit to 1022.1 AUB chkg (negative float)
      'salary'  → expected debit to 9011.1 Direct Salaries (positive float)
      'output2' → expected debit for 9012.6 Employee Payroll Taxes (positive float)
      'output3' → expected debit for 9012.2 401K Match (positive float)
      'output4' → expected debit for 9012.4 Benefits Fees-Admin (positive float)
    """
    sheet_name = None
    for candidate in ("Pivot Check", "Pivot Check JVV"):
        if candidate in wb.sheetnames:
            sheet_name = candidate
            break
    if sheet_name is None:
        return {}

    ws = wb[sheet_name]
    expected = {}

    for row in ws.iter_rows(min_row=1, max_row=100):
        row_map = {cell.column_letter: cell.value for cell in row if cell.value is not None}
        h_raw = row_map.get('H')
        if h_raw is None:
            continue
        try:
            h = round(float(h_raw), 1)
        except (TypeError, ValueError):
            continue

        j_val = row_map.get('J')
        k_val = row_map.get('K')

        def _j(v):
            return float(v) if isinstance(v, (int, float)) and v else None
        def _k(v):
            return float(v) if isinstance(v, (int, float)) and v else None

        if h == 1022.1 and k_val is not None and 'cash' not in expected:
            # AUB checking account — credit (negative) = cash that left the bank
            expected['cash'] = _k(k_val)
        elif h == 9011.1 and j_val is not None and 'salary' not in expected:
            # Direct Salaries total
            expected['salary'] = _j(j_val)
        elif h == 9012.2 and j_val is not None and 'output3' not in expected:
            # 401K Match expense total
            expected['output3'] = _j(j_val)
        elif h == 9012.4 and j_val is not None and 'output4' not in expected:
            # Benefits Fees-Admin total
            expected['output4'] = _j(j_val)
        elif h == 9012.6 and j_val is not None and 'output2' not in expected:
            # Employee Payroll Taxes total (take first J value = employee ER expense)
            expected['output2'] = _j(j_val)

    return expected


def validate_against_pivot_check(cash_amount, salary_total,
                                 o2_total, o3_total, o4_total,
                                 expected, tolerance=0.02):
    """
    Compare script-computed amounts against the Pivot Check expected values.
    Returns True only if all amounts match within tolerance.

    A mismatch almost always means a labor code in the source file has no
    class mapping — either the code is missing from the Reference Class Table,
    or it is an encoding error in the Paycom export (e.g. '3.3.4.1' instead
    of a valid 5-digit code).  The IIF must NOT be imported until the source
    is corrected and re-exported.
    """
    checks = [
        ("Cash 1022.1 AUB chkg",          cash_amount,  expected.get('cash')),
        ("Salary 9011.1",                  salary_total, expected.get('salary')),
        ("Taxes 9012.6 (Output 2 total)",  o2_total,     expected.get('output2')),
        ("401K 9012.2  (Output 3 total)",  o3_total,     expected.get('output3')),
        ("Fees 9012.4  (Output 4 total)",  o4_total,     expected.get('output4')),
    ]
    all_ok = True
    for label, actual, exp_val in checks:
        if actual is None or exp_val is None:
            print(f"  ─ {label}: no expected value in Pivot Check, skipping")
            continue
        diff = abs(actual - exp_val)
        if diff > tolerance:
            print(f"  ❌ MISMATCH {label}: computed={actual:.2f}, "
                  f"Pivot Check={exp_val:.2f}, diff={diff:.2f}")
            all_ok = False
        else:
            print(f"  ✓ {label}: {actual:.2f}")
    if not all_ok:
        print("\n  ⚠  PIVOT CHECK VALIDATION FAILED — do NOT import.")
        print("  Likely cause: a labor code in the Paycom source has no class mapping,")
        print("  or is an encoding error (e.g. '3.3.4.1' instead of a valid 5-digit code).")
        print("  Check the WARNING lines above, correct the source in Paycom, and re-export.")
    return all_ok


def read_pivot_check_adjustments(wb):
    """DEPRECATED — use read_pivot_check_all() instead."""
    return []


# ---------------------------------------------------------------------------
# Output 1 Audit File (Excel)
# ---------------------------------------------------------------------------

def write_output1_audit(audit_path, pay_dt, mp,
                        salary_total,
                        oss_raw, noncash_adjs, oss,
                        pivot_adjustments,
                        balance_entries,
                        cell, exp, mil,
                        output2_total, output3_total, output4_total,
                        garn, ee_401k, fsa, dental, legal, llc_tuition,
                        health, vision, dis_ad, dis_ltd,
                        cash_amount, net):
    """
    Write an Excel audit file for Output 1.
    Captures every intermediate value used to build the JE so that
    discrepancies can be identified before importing into QuickBooks.
    The file is saved alongside the IIF with the suffix ' Audit.xlsx'.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("  WARNING: openpyxl not available — audit file skipped.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Output 1 Audit"

    # ---- Style helpers ----
    def _hfont(bold=False, size=10, italic=False, color="000000"):
        return Font(name="Arial", bold=bold, size=size, italic=italic, color=color)

    SEC_FILL = PatternFill("solid", start_color="D9E1F2")   # blue-grey header
    OK_FILL  = PatternFill("solid", start_color="C6EFCE")   # green
    ERR_FILL = PatternFill("solid", start_color="FFC7CE")   # red
    NUM_FMT  = '#,##0.00;[Red](#,##0.00);"-"'

    r = [1]   # mutable row pointer

    def _next():
        r[0] += 1

    def _blank():
        r[0] += 1

    def _section(title):
        ws.merge_cells(start_row=r[0], start_column=1, end_row=r[0], end_column=4)
        c = ws.cell(row=r[0], column=1, value=title)
        c.font = _hfont(bold=True, size=11)
        c.fill = SEC_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r[0]].height = 18
        _next()

    def _row(label, amount, note="", indent=False):
        prefix = "    " if indent else ""
        ws.cell(row=r[0], column=1, value=prefix + label).font = _hfont()
        c = ws.cell(row=r[0], column=2, value=amount)
        c.number_format = NUM_FMT
        c.font = _hfont()
        c.alignment = Alignment(horizontal="right")
        if note:
            n = ws.cell(row=r[0], column=3, value=note)
            n.font = _hfont(italic=True, color="595959")
        _next()

    def _total_row(label, amount, note=""):
        ws.cell(row=r[0], column=1, value=label).font = _hfont(bold=True)
        c = ws.cell(row=r[0], column=2, value=amount)
        c.number_format = NUM_FMT
        c.font = _hfont(bold=True)
        c.alignment = Alignment(horizontal="right")
        if note:
            ws.cell(row=r[0], column=3, value=note).font = _hfont(italic=True, color="595959")
        _next()

    # ============================================================
    # TITLE
    # ============================================================
    ws.merge_cells(start_row=r[0], start_column=1, end_row=r[0], end_column=4)
    t = ws.cell(row=r[0], column=1, value="OUTPUT 1 PAYROLL JE — BALANCE AUDIT")
    t.font = _hfont(bold=True, size=14)
    _next()

    ws.merge_cells(start_row=r[0], start_column=1, end_row=r[0], end_column=4)
    ws.cell(row=r[0], column=1,
            value=f"Pay Date: {pay_dt.strftime('%m/%d/%Y')}   |   Period: {mp}").font = _hfont(size=10, italic=True)
    _next()

    ws.merge_cells(start_row=r[0], start_column=1, end_row=r[0], end_column=4)
    ws.cell(row=r[0], column=1,
            value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = _hfont(size=9, italic=True, color="888888")
    _next()
    _blank()

    # ============================================================
    # 1. SALARY
    # ============================================================
    _section("1.  SALARY DEBITS")
    _total_row("Total Salary Debits (Output 1 sheet)", salary_total,
               note="Sum of all class-coded salary rows")
    _blank()

    # ============================================================
    # 2. OSS — MANUAL PAYMENTS
    # ============================================================
    _section("2.  OSS — MANUAL PAYMENTS (Payroll Suspense Clearing)")
    _total_row("OSS  (Pivot Check — Name: OSS row)", oss,
               note="Net cash value set by operator in Pivot Check; no Sheet0 adjustment needed")
    _blank()

    # ============================================================
    # 3. PIVOT CHECK / OTHER MANUAL ADJUSTMENTS
    # ============================================================
    _PRIOR_KW = ('manual check', 'advance payment', 'advance paid',
                 'manual ck', 'prior payment')

    _section("3.  PIVOT CHECK / OTHER MANUAL ADJUSTMENTS (each gets its own SPL)")
    if pivot_adjustments:
        for adj in pivot_adjustments:
            is_prior = any(kw in (adj.get('memo') or '').lower() for kw in _PRIOR_KW)
            note = "prior-period (also removed from OSS above)" if is_prior else ""
            _row(adj['name'], adj['amount'], note=note)
        _total_row("Total Pivot Adjustments",
                   sum(adj['amount'] for adj in pivot_adjustments))
    else:
        ws.cell(row=r[0], column=1, value="  (none)").font = _hfont(italic=True, color="888888")
        _next()
    _blank()

    # ============================================================
    # 4. BALANCE ENTRIES FROM SHEET0
    # ============================================================
    _section("4.  BALANCE ENTRIES — SHEET0 AGGREGATES")
    # Debits (positive in JE)
    _row("Cell Phone Reimbursement  (debit)", abs(cell),  note="6421")
    _row("Expense Reimbursement     (debit)", abs(exp),   note="6422")
    _row("Mileage Reimbursement     (debit)", abs(mil),   note="6420")
    # Credits (negative in JE)
    _row("Garnishment Withheld               (credit)", garn,       note="6412")
    _row("EE 401K Contributions Payable      (credit)", ee_401k,    note="6413")
    _row("FSA Withheld                       (credit)", fsa,        note="6417")
    _row("Dental Premiums — MetLife          (credit)", dental,     note="6419")
    _row("Vision Premiums — UHC              (credit)", vision,     note="6419")
    _row("Legal Resources Withheld           (credit)", legal,      note="6423")
    _row("LLC Tuition Fees Withheld          (credit)", llc_tuition, note="6415")
    _row("Health — Sentara                   (credit)", health,     note="9012.1")
    _row("Disability & Life — AD&D           (credit)", dis_ad,     note="9012.3")
    _row("Disability — LTD                   (credit)", dis_ltd,    note="9012.3")
    _blank()

    # ============================================================
    # 5. CROSS-OUTPUT TOTALS
    # ============================================================
    _section("5.  CROSS-OUTPUT TOTALS")
    _row("Output 2 Total — Payroll Taxes     (debit)",         output2_total)
    _row("Output 3 Total — 401K Match        (debit + credit offset)", output3_total)
    _row("Output 4 Total — Paycom Fee        (debit)",         output4_total)
    _blank()

    # ============================================================
    # 6. CASH PLUG CALCULATION
    # ============================================================
    _section("6.  CASH PLUG CALCULATION")
    debit_items_   = [abs(cell), abs(exp), abs(mil), output3_total, output4_total, output2_total]
    pivot_adj_net_ = sum(adj['amount'] for adj in (pivot_adjustments or []))
    credit_items_  = [garn, ee_401k, -output3_total, fsa, dental, legal, llc_tuition,
                      health, vision, dis_ad, dis_ltd, oss, pivot_adj_net_]
    total_debits_  = salary_total + sum(debit_items_)
    total_credits_ = sum(c for c in credit_items_ if c < 0)

    _row("Total Salary Debits",                                  salary_total)
    _row("+ Reimbursements (Cell Phone + Expense + Mileage)",    sum([abs(cell), abs(exp), abs(mil)]))
    _row("+ Output 3 Match Debit",                               output3_total)
    _row("+ Output 4 Fee Debit",                                 output4_total)
    _row("+ Output 2 Tax Debit",                                 output2_total)
    _total_row("= Total Debits",                                 total_debits_)
    _blank()
    _row("Total Credits (negative items only)",                  total_credits_)
    _total_row("Cash Plug = −(Total Debits + Total Credits)",    cash_amount,
               note="This is the ACH / cash SPL line in the JE")
    _blank()

    # ============================================================
    # 7. BALANCE VERIFICATION
    # ============================================================
    _section("7.  BALANCE VERIFICATION")
    ws.cell(row=r[0], column=1,
            value="Net of all amounts (must equal 0.00)").font = _hfont(bold=True, size=11)
    c = ws.cell(row=r[0], column=2, value=net)
    c.number_format = NUM_FMT
    c.font = _hfont(bold=True, size=11)
    c.alignment = Alignment(horizontal="right")
    balanced = abs(net) < 0.02
    c.fill = OK_FILL if balanced else ERR_FILL
    note_val = "✓  BALANCED — safe to import" if balanced else f"⚠  OUT OF BALANCE by {net:,.2f} — DO NOT IMPORT"
    note_color = "375623" if balanced else "9C0006"
    n = ws.cell(row=r[0], column=3, value=note_val)
    n.font = _hfont(bold=True, color=note_color)
    _next()
    _blank()

    # ============================================================
    # COLUMN WIDTHS & FREEZE
    # ============================================================
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 52
    ws.freeze_panes = "A6"

    wb.save(audit_path)
    print(f"  ✓ Output 1 audit file: {os.path.basename(audit_path)}")


# ---------------------------------------------------------------------------
# IIF generators
# ---------------------------------------------------------------------------

def generate_simple_output(rows, class_table, account, docnum, memo_full, date_str, output_path):
    """
    Generate Outputs 2, 3, or 4.
    Debits each class with its amount; final SPL credits 901 Unallocated for negative sum.
    """
    iif_rows = []
    total = 0.0
    first = True

    for code, amount in rows:
        qb_class = lookup_class(code, class_table)
        if qb_class is None:
            print(f"  WARNING: No class mapping for code '{code}' — skipping.")
            continue

        total += amount
        row_type = "TRNS" if first else "SPL"
        doc = docnum if first else ""
        iif_rows.append(build_row(
            row_type, "GENERAL JOURNAL", date_str,
            account, "", qb_class, amount, doc, memo_full
        ))
        first = False

    # Final credit row: 901 Unallocated, negative of total
    iif_rows.append(build_row(
        "SPL", "GENERAL JOURNAL", date_str,
        account, "", CLASS_UNALLOCATED, -total, "", memo_full
    ))

    write_iif(output_path, build_iif_header(), iif_rows)
    return total


def generate_output1(salary_rows, class_table, balance_entries,
                     output2_total, output3_total, output4_total,
                     docnum, memo_prefix_str, date_str, output_path,
                     prior_pay_date="", pay_dt=None,
                     pivot_adjustments=None):
    """
    Generate Output 1 — Payroll JE.
    salary_rows: list of (code, amount) for the salary debit section.
    balance_entries: dict from aggregate_balance_entries().
    output2/3/4_total: totals from the other three outputs.
    """
    iif_rows = []
    first = True

    # 1. Salary debit lines (by class)
    for code, amount in salary_rows:
        qb_class = lookup_class(code, class_table)
        if qb_class is None:
            print(f"  WARNING: No class mapping for salary code '{code}' — skipping.")
            continue

        row_type = "TRNS" if first else "SPL"
        doc = docnum if first else ""
        memo = f"{memo_prefix_str} 9011.1 Direct Salaries (Full Time)"
        iif_rows.append(build_row(
            row_type, "GENERAL JOURNAL", date_str,
            ACCT_SALARY, "", qb_class, amount, doc, memo
        ))
        first = False

    # Compute total salary debits
    salary_total = sum(r[1] for r in salary_rows
                       if lookup_class(r[0], class_table) is not None)

    # 2. Build balance/offset entries
    # Determine memos
    mp = memo_prefix_str

    # All balance SPL rows — build them in the canonical order
    # (Cash is the plug computed last)

    balance_spls = []

    def spl(accnt, name, cls, amount, memo):
        balance_spls.append(build_row(
            "SPL", "GENERAL JOURNAL", date_str,
            accnt, name, cls, amount, "", memo
        ))

    # Payroll Suspense (always 0)
    spl(ACCT_SUSPENSE, " ", " ", 0.00, " ")

    # OSS is read directly from the Pivot Check (the "Name: OSS" row in the 6410 section).
    # The Pivot Check operator has already set the correct net OSS value — no adjustment needed.
    noncash_adjs = []   # no longer used; retained so audit call below stays compatible
    oss_raw = balance_entries.get("oss_manual", 0.0)
    oss     = oss_raw   # Pivot Check value is already correct
    if oss != 0.0:
        print(f"  OSS (from Pivot Check): {oss:,.2f}")

    # Manual check payments (OSS) — only current-period paper checks
    if oss != 0.0:
        ppe = f"{pay_dt.month}.{pay_dt.day:02d}.{str(pay_dt.year)[2:]}"
        spl(ACCT_SUSPENSE_MANUAL, "", "", oss,
            f"{mp} Paycom OSS cash payments for PPE {ppe}")

    # Additional 6410 manual adjustments (void checks, negative leave repay, etc.)
    # These come from the Pivot Check sheet and are not in Sheet0.
    for adj in (pivot_adjustments or []):
        # Strip "OSS" from the name field — it is internal Paycom terminology
        # and should not appear in the QuickBooks IIF name column.
        iif_name = "" if (adj['name'] or "").strip().upper() == "OSS" else adj['name']
        spl(ACCT_SUSPENSE_MANUAL, iif_name, "", adj['amount'], f"{mp} {adj['memo']}")

    # Garnishment Withheld
    garn = balance_entries.get("garnishment", 0.0)
    if garn != 0.0:
        spl(ACCT_GARNISHMENT, "", " ", garn, f"{mp} 6412 Garnishment withheld")

    # Cell Phone Reimbursement (positive debit)
    cell = balance_entries.get("cell_phone", 0.0)
    if cell != 0.0:
        spl(ACCT_CELL_PHONE, "", "", abs(cell), f"{mp} 6421 Cell Phone Reimbursement")

    # Expense Reimbursement (positive debit)
    exp = balance_entries.get("expense_reimb", 0.0)
    if exp != 0.0:
        spl(ACCT_EXPENSE_REIMB, "", " ", abs(exp), f"{mp} 6422 Expense Reimbursement")

    # Mileage Reimbursement (positive debit)
    mil = balance_entries.get("mileage", 0.0)
    if mil != 0.0:
        spl(ACCT_MILEAGE, "", "", abs(mil), f"{mp} 6420 Mileage Reimbursement")

    # 401K Contributions Payable — EE (credit, negative)
    ee_401k = balance_entries.get("ee_401k", 0.0)
    if ee_401k != 0.0:
        spl(ACCT_401K_PAYABLE, "", "", ee_401k,
            f"{mp} 6413  401K Contributions Payable - Employee contributions")

    # 401K Contributions Payable — Match (credit, negative, = Output 3 total)
    spl(ACCT_401K_PAYABLE, "", "", -output3_total,
        f"{mp} 6413  401K Contributions Payable - Match")

    # FSA Withheld (credit, negative)
    fsa = balance_entries.get("fsa", 0.0)
    if fsa != 0.0:
        spl(ACCT_FSA, "", "", fsa, f"{mp} 6417  FSA Withheld")

    # Dental Premiums — MetLife (credit, negative)
    dental = balance_entries.get("dental_metlife", 0.0)
    if dental != 0.0:
        spl(ACCT_DENTAL, "MetLife Small Business Center", "", dental,
            f"{mp} 6419  Dental Premiums Withheld")

    # Legal Resources Withheld (credit, negative)
    legal = balance_entries.get("legal", 0.0)
    if legal != 0.0:
        spl(ACCT_LEGAL, "", " ", legal, f"{mp} 6423  Legal Resources Withheld")

    # LLC Tuition Fees Withheld (credit, negative) — 6415
    llc_tuition = balance_entries.get("llc_tuition", 0.0)
    if llc_tuition != 0.0:
        spl(ACCT_LLC_TUITION, "", "", llc_tuition, f"{mp} 6415  LLC Tuition Fees Withheld")

    # Health — Sentara (class 901, credit, negative)
    health = balance_entries.get("health_sentara", 0.0)
    if health != 0.0:
        spl(ACCT_HEALTH, "Sentara Healthcare", CLASS_UNALLOCATED, health,
            f"{mp} 9012.1  Health - Sentara")

    # Vision / Dental — UHC (credit, negative)
    vision = balance_entries.get("vision_uhc", 0.0)
    if vision != 0.0:
        spl(ACCT_DENTAL, "United Healthcare", "", vision,
            f"{mp} 6419  Vision Premiums Withheld")

    # 401K fringe expense (class 901, debit = Output 3 total, positive)
    spl(ACCT_401K, "", CLASS_UNALLOCATED, output3_total,
        f"{mp} 9012.2  401K Match")

    # Disability & Life — Life & AD&D (class 901, credit, negative)
    dis_ad = balance_entries.get("disability_ad", 0.0)
    if dis_ad != 0.0:
        spl(ACCT_DISABILITY, "", CLASS_UNALLOCATED, dis_ad,
            f"{mp} 9012.3  Disability & Life Ins. - Life & AD&D")

    # Disability — LTD (class 901, credit, negative)
    dis_ltd = balance_entries.get("disability_ltd", 0.0)
    if dis_ltd != 0.0:
        spl(ACCT_DISABILITY, "", CLASS_UNALLOCATED, dis_ltd,
            f"{mp} 9012.3  Disability & Life Ins. - LTD")

    # Benefits Fees-Admin (class 901, debit = Output 4 total, positive)
    spl(ACCT_FEES, "", CLASS_UNALLOCATED, output4_total,
        f"{mp} 9012.4  Benefits Fees-Admin")

    # Employee Payroll Taxes (class 901, debit = Output 2 total, positive)
    spl(ACCT_TAX, "", CLASS_UNALLOCATED, output2_total,
        f"{mp} 9012.6  Employee Payroll Taxes")

    # 3. Compute the cash plug (balancing amount)
    # Sum of all debits in balance section
    debit_items = [abs(cell), abs(exp), abs(mil), output3_total, output4_total, output2_total]
    # oss is already adjusted above (prior-period amounts removed).
    # ALL pivot adjustments go into credit_items — they each have SPL entries to Suspense.
    pivot_adj_net = sum(adj['amount'] for adj in (pivot_adjustments or []))
    credit_items = [garn, ee_401k, -output3_total, fsa, dental, legal, llc_tuition, health, vision,
                    dis_ad, dis_ltd, oss, pivot_adj_net]

    balance_debits  = salary_total + sum(debit_items)
    balance_credits = sum(c for c in credit_items if c < 0)
    cash_amount     = -(balance_debits + balance_credits)

    # Prepend cash row (comes first in the balance section, before Suspense)
    # Cash memo uses the current pay date by default; --prior-pay-date overrides if needed.
    memo_date = prior_pay_date if prior_pay_date else date_str
    cash_memo = f"{mp} Payment {memo_date} payroll"
    cash_row  = build_row(
        "SPL", "GENERAL JOURNAL", date_str,
        ACCT_CASH, "", " ", cash_amount, "", cash_memo
    )
    # Insert cash row at position 0 of balance_spls (before suspense)
    all_spls = [cash_row] + balance_spls

    iif_rows.extend(all_spls)

    write_iif(output_path, build_iif_header(), iif_rows)

    # Balance check
    all_amounts = []
    for code, amount in salary_rows:
        if lookup_class(code, class_table):
            all_amounts.append(amount)
    all_amounts.extend(debit_items)
    all_amounts.extend(credit_items)
    all_amounts.append(cash_amount)
    all_amounts.append(0.0)  # suspense

    net = sum(all_amounts)
    if abs(net) > 1.0:
        print(f"  ⚠ WARNING: Output 1 does not balance! Net = {net:.2f}")
    else:
        print(f"  ✓ Output 1 balance check: net = {net:.2f}")

    # Write Excel audit file to the local output folder (NOT the Drive Outbox)
    # so it never appears alongside the IIF files in the outbox.
    _script_dir   = os.path.dirname(os.path.abspath(__file__))
    _local_output = os.path.join(_script_dir, "output files")
    os.makedirs(_local_output, exist_ok=True)
    _audit_name = os.path.basename(output_path).replace(".iif", " Audit.xlsx")
    audit_path  = os.path.join(_local_output, _audit_name)
    write_output1_audit(
        audit_path, pay_dt, memo_prefix_str,
        salary_total,
        oss_raw, noncash_adjs, oss,
        pivot_adjustments,
        balance_entries,
        cell, exp, mil,
        output2_total, output3_total, output4_total,
        garn, ee_401k, fsa, dental, legal, llc_tuition,
        health, vision, dis_ad, dis_ltd,
        cash_amount, net,
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Generate Cornerstones payroll IIF files.")
    parser.add_argument("--source",      required=True,  help="Path to Source 1 xlsx")
    parser.add_argument("--class-ref",   required=False, help="Path to Reference 1 Class Table xlsx")
    parser.add_argument("--pay-date",    required=True,  help="Pay date, e.g. '2/13/2026'")
    parser.add_argument("--pay-period",  required=True,  type=int, help="Pay period number, e.g. 17")
    parser.add_argument("--fiscal-year", required=True,  type=int, help="2-digit fiscal year, e.g. 26")
    parser.add_argument("--output-dir",  required=False, default=".", help="Output directory")
    parser.add_argument("--prior-pay-date", required=False, default="",
                        help="Prior pay date for cash line memo (e.g. '1/30/2026'). "
                             "If omitted, memo reads 'FY##.PR.## Payment payroll'.")
    parser.add_argument("--extra-class", action="append", default=[], metavar="CODE=QB_PATH",
                        help="Add or override a class mapping. May be repeated. "
                             "Example: --extra-class '32112=30000 Housing ...:32112 Exp Op...'")
    parser.add_argument("--class-final-check", required=False, default="",
                        help="Path to 'Reference 2 Class Table FINAL CHECK.xlsx'. "
                             "When supplied, all CLASS values in the generated IIF files are "
                             "verified against this table. Any missing class causes the script "
                             "to exit with an error after listing the offending values.")
    args = parser.parse_args()

    pay_dt    = parse_pay_date(args.pay_date)
    date_str  = iif_date(pay_dt)
    label     = file_date_label(pay_dt)   # e.g. "13 Feb"
    mp        = memo_prefix(args.fiscal_year, args.pay_period)

    print(f"\n=== Cornerstones Payroll IIF Generator ===")
    print(f"Pay Date:   {date_str}")
    print(f"Pay Period: {mp}")
    print(f"Output Dir: {args.output_dir}")
    print()

    os.makedirs(args.output_dir, exist_ok=True)

    # Load source workbook
    print("Loading source workbook...")
    wb = openpyxl.load_workbook(args.source, read_only=True, data_only=True)

    # Load class table
    print("Building class lookup table...")
    class_table = load_class_table(wb, args.class_ref)

    # Apply any extra class overrides from command line
    if args.extra_class:
        for entry in args.extra_class:
            if "=" in entry:
                code, path = entry.split("=", 1)
                class_table[code.strip()] = path.strip()
                print(f"  Extra class override: {code.strip()} → {path.strip()[:60]}")
            else:
                print(f"  WARNING: --extra-class '{entry}' ignored (expected CODE=QB_PATH format)")

    # Read Output sheets — try multiple possible sheet name variants
    def find_sheet(wb, *candidates):
        """Return the first sheet name from candidates that exists in the workbook."""
        for name in candidates:
            if name in wb.sheetnames:
                return name
        raise KeyError(f"None of the expected sheets found: {candidates}. "
                       f"Available: {wb.sheetnames}")

    print("\nReading Output sheets...")
    o1_rows = read_output_sheet(wb, find_sheet(wb, "Output 1 Salary By Code", "Salary By Code"), class_table)
    o2_rows = read_output_sheet(wb, find_sheet(wb, "Output 2 Taxes By Code",  "Taxes By Code"),  class_table)
    o3_rows = read_output_sheet(wb, find_sheet(wb, "Output 3 401k Fringe By Code", "Fringe By Code", "401k Fringe By Code"), class_table)
    o4_rows = read_output_sheet(wb, find_sheet(wb, "Output 4 Paycom Fee By Code", "Paycom Fee By Code"), class_table)
    print(f"  Output 1 rows: {len(o1_rows)}")
    print(f"  Output 2 rows: {len(o2_rows)}")
    print(f"  Output 3 rows: {len(o3_rows)}")
    print(f"  Output 4 rows: {len(o4_rows)}")

    # Read ALL Output 1 balance entries from the Pivot Check sheet (columns G–N).
    # Sheet0 is no longer used — the Pivot Check is the single source of truth.
    print("\nReading Pivot Check for all Output 1 balance entries...")
    balance_entries, pivot_adjustments = read_pivot_check_all(wb)
    print(f"  OSS:            {balance_entries['oss_manual']:.2f}")
    print(f"  Garnishment:    {balance_entries['garnishment']:.2f}")
    print(f"  Cell Phone:     {balance_entries['cell_phone']:.2f}")
    print(f"  Expense Reimb:  {balance_entries['expense_reimb']:.2f}")
    print(f"  Mileage:        {balance_entries['mileage']:.2f}")
    print(f"  EE 401K:        {balance_entries['ee_401k']:.2f}")
    print(f"  FSA:            {balance_entries['fsa']:.2f}")
    print(f"  Dental MetLife: {balance_entries['dental_metlife']:.2f}")
    print(f"  Vision UHC:     {balance_entries['vision_uhc']:.2f}")
    print(f"  Legal:          {balance_entries['legal']:.2f}")
    print(f"  Health Sentara: {balance_entries['health_sentara']:.2f}")
    print(f"  Disability AD:  {balance_entries['disability_ad']:.2f}")
    print(f"  Disability LTD: {balance_entries['disability_ltd']:.2f}")
    print(f"  LLC Tuition:    {balance_entries['llc_tuition']:.2f}")
    if not pivot_adjustments:
        print("  (no individual 6410 adjustments found)")

    # Read Pivot Check expected amounts for final validation
    print("\nReading Pivot Check expected amounts for validation...")
    pivot_expected = read_pivot_check_expected_amounts(wb)
    if pivot_expected:
        for k, v in pivot_expected.items():
            print(f"  Pivot Check expects  {k}: {v:.2f}")
    else:
        print("  (Pivot Check not found — validation will be skipped)")

    wb.close()

    # ---- Generate Output 2 (Taxes) ----
    print("\nGenerating Output 2 — Taxes JE...")
    o2_path  = os.path.join(args.output_dir, f"Output 2 Taxes JE {label}.iif")
    o2_total = generate_simple_output(
        o2_rows, class_table, ACCT_TAX,
        docnum_tax(pay_dt),
        f"{mp} 9012.6   Employee Payroll Taxes",
        date_str, o2_path
    )
    print(f"  Total: {o2_total:.2f}")

    # ---- Generate Output 3 (401K) ----
    print("\nGenerating Output 3 — 401K JE...")
    o3_path  = os.path.join(args.output_dir, f"Output 3 Payroll 401k JE {label}.iif")
    o3_total = generate_simple_output(
        o3_rows, class_table, ACCT_401K,
        docnum_401k(pay_dt),
        f"{mp} 9012.2 401K Match",
        date_str, o3_path
    )
    print(f"  Total: {o3_total:.2f}")

    # ---- Generate Output 4 (Payroll Fee) ----
    print("\nGenerating Output 4 — Payroll Fee JE...")
    o4_path  = os.path.join(args.output_dir, f"Output 4 Payroll Fee JE {label}.iif")
    o4_total = generate_simple_output(
        o4_rows, class_table, ACCT_FEES,
        docnum_fee(pay_dt),
        f"{mp} 9012.4   Benefits Fees-Admin",
        date_str, o4_path
    )
    print(f"  Total: {o4_total:.2f}")

    # ---- Generate Output 1 (Payroll JE) ----
    print("\nGenerating Output 1 — Payroll JE...")
    o1_path = os.path.join(args.output_dir, f"Output 1 Payroll JE {label}.iif")
    generate_output1(
        o1_rows, class_table, balance_entries,
        o2_total, o3_total, o4_total,
        docnum_prp(pay_dt), mp, date_str, o1_path,
        prior_pay_date=args.prior_pay_date,
        pay_dt=pay_dt,
        pivot_adjustments=pivot_adjustments
    )

    # ---- Pivot Check amount validation ----
    # Compute the salary total that was actually written (net amounts after Cr offsets)
    salary_total_written = sum(
        net for code, net in o1_rows
        if lookup_class(code, class_table) is not None
    )
    print("\n=== Pivot Check Amount Validation ===")
    pc_ok = validate_against_pivot_check(
        cash_amount=None,          # fetched from the generated file below
        salary_total=salary_total_written,
        o2_total=o2_total,
        o3_total=o3_total,
        o4_total=o4_total,
        expected=pivot_expected,
    )
    # Also check the cash line from the generated Output 1 file
    if pivot_expected.get('cash') is not None:
        cash_written = None
        try:
            with open(o1_path, 'r', encoding='utf-8') as fh:
                for line in fh:
                    if 'Atlantic Union' in line:
                        parts = line.rstrip('\n\r').split('\t')
                        cash_written = float(parts[7]) if len(parts) > 7 else None
                        break
        except Exception:
            pass
        if cash_written is not None:
            exp_cash = pivot_expected['cash']
            diff = abs(cash_written - exp_cash)
            if diff > 0.02:
                print(f"  ❌ MISMATCH Cash 1022.1 AUB chkg: "
                      f"written={cash_written:.2f}, Pivot Check={exp_cash:.2f}, diff={diff:.2f}")
                pc_ok = False
            else:
                print(f"  ✓ Cash 1022.1 AUB chkg: {cash_written:.2f}")
    if pc_ok:
        print("  ✓ All Pivot Check amounts match — safe to import.")
    else:
        print("\n  ❌ PIVOT CHECK VALIDATION FAILED — do NOT import these files.")
        print("  Common causes:")
        print("    1. An output sheet has a Cr column (col E) offset that was not netted.")
        print("    2. A withholding (e.g., 6415 LLC Tuition) is missing from balance entries.")
        print("    3. A Pivot Check 6410 manual adjustment was double-counted or omitted.")
        sys.exit(1)

    # ---- Final class verification against Reference 2 ----
    if args.class_final_check:
        print("\nRunning final class verification against Reference 2...")
        final_check_table = load_final_check_table(args.class_final_check)
        print(f"  Loaded {len(final_check_table)} valid class paths from Reference 2.")
        missing = verify_classes_against_final_check(
            [o1_path, o2_path, o3_path, o4_path], final_check_table
        )
        if missing:
            print("\n  ❌ FINAL CHECK FAILED — the following class values are not in Reference 2:")
            for fname, cls in missing:
                print(f"     [{fname}]  '{cls}'")
            print("\n  Process stopped. Fix the class mappings and regenerate before importing.")
            sys.exit(1)
        else:
            print("  ✓ All class values verified against Reference 2.")

    print("\n=== Done! ===")
    print(f"Files saved to: {args.output_dir}")


def run_programmatic(source_path, class_ref_path, class_final_check_path,
                     pay_date_str, pay_period, fiscal_year, output_dir,
                     prior_pay_date_str="", extra_classes=None):
    """
    Call the generator programmatically (used by the Streamlit app).

    Returns:
        success (bool), log (str), output_files (dict[filename -> str content])
    """
    import io
    import contextlib

    buf = io.StringIO()
    output_files = {}
    success = False

    def _run():
        nonlocal success
        pay_dt   = parse_pay_date(pay_date_str)
        date_str = iif_date(pay_dt)
        label    = file_date_label(pay_dt)
        mp       = memo_prefix(fiscal_year, pay_period)

        print(f"\n=== Cornerstones Payroll IIF Generator ===")
        print(f"Pay Date:   {date_str}")
        print(f"Pay Period: {mp}")
        print(f"Output Dir: {output_dir}\n")

        os.makedirs(output_dir, exist_ok=True)

        print("Loading source workbook...")
        wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)

        print("Building class lookup table...")
        class_table = load_class_table(wb, class_ref_path)

        if extra_classes:
            for entry in extra_classes:
                if "=" in entry:
                    code, path = entry.split("=", 1)
                    class_table[code.strip()] = path.strip()
                    print(f"  Extra class override: {code.strip()} → {path.strip()[:60]}")

        def find_sheet(wb, *candidates):
            for name in candidates:
                if name in wb.sheetnames:
                    return name
            raise KeyError(f"None of the expected sheets found: {candidates}. "
                           f"Available: {wb.sheetnames}")

        print("\nReading Output sheets...")
        o1_rows = read_output_sheet(wb, find_sheet(wb, "Output 1 Salary By Code", "Salary By Code"), class_table)
        o2_rows = read_output_sheet(wb, find_sheet(wb, "Output 2 Taxes By Code",  "Taxes By Code"),  class_table)
        o3_rows = read_output_sheet(wb, find_sheet(wb, "Output 3 401k Fringe By Code", "Fringe By Code", "401k Fringe By Code"), class_table)
        o4_rows = read_output_sheet(wb, find_sheet(wb, "Output 4 Paycom Fee By Code", "Paycom Fee By Code"), class_table)
        print(f"  Output 1 rows: {len(o1_rows)}")
        print(f"  Output 2 rows: {len(o2_rows)}")
        print(f"  Output 3 rows: {len(o3_rows)}")
        print(f"  Output 4 rows: {len(o4_rows)}")

        print("\nReading Pivot Check for all Output 1 balance entries...")
        balance_entries, pivot_adjustments = read_pivot_check_all(wb)
        if not pivot_adjustments:
            print("  (no individual 6410 adjustments found)")

        print("\nReading Pivot Check expected amounts for validation...")
        pivot_expected = read_pivot_check_expected_amounts(wb)
        if pivot_expected:
            for k, v in pivot_expected.items():
                print(f"  Pivot Check expects  {k}: {v:.2f}")
        else:
            print("  (Pivot Check not found — validation will be skipped)")

        wb.close()

        print("\nGenerating Output 2 — Taxes JE...")
        o2_path  = os.path.join(output_dir, f"Output 2 Taxes JE {label}.iif")
        o2_total = generate_simple_output(o2_rows, class_table, ACCT_TAX,
                       docnum_tax(pay_dt), f"{mp} 9012.6   Employee Payroll Taxes",
                       date_str, o2_path)
        print(f"  Total: {o2_total:.2f}")

        print("\nGenerating Output 3 — 401K JE...")
        o3_path  = os.path.join(output_dir, f"Output 3 Payroll 401k JE {label}.iif")
        o3_total = generate_simple_output(o3_rows, class_table, ACCT_401K,
                       docnum_401k(pay_dt), f"{mp} 9012.2 401K Match",
                       date_str, o3_path)
        print(f"  Total: {o3_total:.2f}")

        print("\nGenerating Output 4 — Payroll Fee JE...")
        o4_path  = os.path.join(output_dir, f"Output 4 Payroll Fee JE {label}.iif")
        o4_total = generate_simple_output(o4_rows, class_table, ACCT_FEES,
                       docnum_fee(pay_dt), f"{mp} 9012.4   Benefits Fees-Admin",
                       date_str, o4_path)
        print(f"  Total: {o4_total:.2f}")

        print("\nGenerating Output 1 — Payroll JE...")
        o1_path = os.path.join(output_dir, f"Output 1 Payroll JE {label}.iif")
        generate_output1(o1_rows, class_table, balance_entries,
                         o2_total, o3_total, o4_total,
                         docnum_prp(pay_dt), mp, date_str, o1_path,
                         prior_pay_date=prior_pay_date_str, pay_dt=pay_dt,
                         pivot_adjustments=pivot_adjustments)

        # Pivot Check validation
        salary_total_written = sum(
            net for code, net in o1_rows
            if lookup_class(code, class_table) is not None
        )
        print("\n=== Pivot Check Amount Validation ===")
        pc_ok = validate_against_pivot_check(
            cash_amount=None,
            salary_total=salary_total_written,
            o2_total=o2_total, o3_total=o3_total, o4_total=o4_total,
            expected=pivot_expected,
        )
        if pivot_expected.get('cash') is not None:
            cash_written = None
            try:
                with open(o1_path, 'r', encoding='utf-8') as fh:
                    for line in fh:
                        if 'Atlantic Union' in line:
                            parts = line.rstrip('\n\r').split('\t')
                            cash_written = float(parts[7]) if len(parts) > 7 else None
                            break
            except Exception:
                pass
            if cash_written is not None:
                exp_cash = pivot_expected['cash']
                diff = abs(cash_written - exp_cash)
                if diff > 0.02:
                    print(f"  ❌ MISMATCH Cash 1022.1: written={cash_written:.2f}, "
                          f"Pivot Check={exp_cash:.2f}, diff={diff:.2f}")
                    pc_ok = False
                else:
                    print(f"  ✓ Cash 1022.1 AUB chkg: {cash_written:.2f}")

        if pc_ok:
            print("  ✓ All Pivot Check amounts match — safe to import.")
        else:
            print("\n  ❌ PIVOT CHECK VALIDATION FAILED — do NOT import.")
            raise RuntimeError("Pivot Check validation failed.")

        # Final class verification
        if class_final_check_path and os.path.exists(class_final_check_path):
            print("\nRunning final class verification against Reference 2...")
            final_check_table = load_final_check_table(class_final_check_path)
            print(f"  Loaded {len(final_check_table)} valid class paths.")
            missing = verify_classes_against_final_check(
                [o1_path, o2_path, o3_path, o4_path], final_check_table)
            if missing:
                print("\n  ❌ FINAL CHECK FAILED — missing class values:")
                for fname, cls in missing:
                    print(f"     [{fname}]  '{cls}'")
                raise RuntimeError("Final class check failed.")
            else:
                print("  ✓ All class values verified against Reference 2.")

        # Collect output files
        for p in [o1_path, o2_path, o3_path, o4_path]:
            if os.path.exists(p):
                with open(p, 'r', encoding='utf-8') as f:
                    output_files[os.path.basename(p)] = f.read()

        print("\n=== Done! ===")
        success = True

    try:
        with contextlib.redirect_stdout(buf):
            _run()
    except (SystemExit, RuntimeError, Exception) as e:
        if not isinstance(e, SystemExit) or e.code != 0:
            buf.write(f"\n\nERROR: {e}\n")
        success = isinstance(e, SystemExit) and e.code == 0

    return success, buf.getvalue(), output_files


if __name__ == "__main__":
    main()
